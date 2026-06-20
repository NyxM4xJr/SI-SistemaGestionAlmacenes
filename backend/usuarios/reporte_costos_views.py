"""
============================================================
ARCHIVO: backend/usuarios/reporte_costos_views.py
CASO DE USO: CU27 - Generar Reporte de Costos por Plato
CICLO: 4
AUTOR: Karen Ortega
FECHA: 19/06/26
============================================================

DESCRIPCIÓN:
Calcula, para cada plato con receta asociada, el costo TEÓRICO
(sin merma) y el costo REAL (con merma técnica de CU22 aplicada)
de producirlo, usando el costo unitario VIGENTE de cada insumo
(el del lote más reciente, según LOTE.created_at). Compara ambos
contra el precio de venta (PLATO.costo) para obtener el margen.

No se crea ninguna tabla nueva. Se generan archivos PDF (reportlab)
y Excel (openpyxl) reales para descarga. Solo la exportación deja
constancia en la bitácora (GENERAR_REPORTE_COSTOS).

Tablas consultadas (todas ya existentes):
- PLATO
- RECETA          (plato_id FK -> PLATO)
- DETALLE_RECETA  (receta_id FK -> RECETA, insumo_id FK -> INSUMO)
- LOTE            (created_at, usado para encontrar el más reciente)
- DETALLE_LOTE    (lote_id FK -> LOTE, insumo_id FK -> INSUMO, costo_unitario)
- FICHA_TECNICA   (insumo_id, porcentaje_merma)
- INSUMO          (para mostrar el nombre, si se requiere detalle)
"""

import io
import logging

from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from supabase import create_client
from django.conf import settings
from django.http import HttpResponse

from bitacora.utils import registrar_accion, obtener_ip_cliente

logger = logging.getLogger(__name__)


def _calcular_reporte_costos(supabase, plato_id=None):
    """
    Función compartida por los 3 endpoints (JSON, PDF, Excel).
    Calcula costo teórico, costo real y margen para los platos
    con receta asociada (o para un solo plato si se indica plato_id).

    Retorna una lista de dicts:
    [
        {
            "plato_id": int,
            "plato_nombre": str,
            "costo_teorico": float,
            "costo_real": float,
            "precio_venta": float | None,
            "margen": float | None,
            "costos_incompletos": bool
        },
        ...
    ]
    """
    # 1) Recetas (una por plato, según el patrón ya establecido en CU21)
    recetas_query = supabase.table('receta').select('id, plato_id')
    if plato_id is not None:
        recetas_query = recetas_query.eq('plato_id', plato_id)
    recetas = recetas_query.execute().data or []

    if not recetas:
        return []

    receta_ids = [r['id'] for r in recetas]
    receta_a_plato = {r['id']: r['plato_id'] for r in recetas}
    plato_ids = list({r['plato_id'] for r in recetas})

    # 2) Detalles de receta (insumos + cantidades)
    detalles = supabase.table('detalle_receta') \
        .select('receta_id, insumo_id, cantidad') \
        .in_('receta_id', receta_ids) \
        .execute().data or []

    insumo_ids = list({d['insumo_id'] for d in detalles})

    # 3) Costo unitario vigente por insumo: el detalle_lote del lote
    #    con created_at más reciente para ese insumo_id.
    costo_unitario_por_insumo = {}
    if insumo_ids:
        detalle_lotes = supabase.table('detalle_lote') \
            .select('insumo_id, costo_unitario, lote:lote_id(created_at)') \
            .in_('insumo_id', insumo_ids) \
            .execute().data or []

        # Para cada insumo, nos quedamos con el detalle_lote cuyo
        # lote.created_at sea el más reciente.
        mas_reciente_por_insumo = {}  # insumo_id -> created_at string
        for dl in detalle_lotes:
            insumo_id = dl['insumo_id']
            lote_info = dl.get('lote') or {}
            created_at = lote_info.get('created_at')
            if not created_at:
                continue
            if (
                insumo_id not in mas_reciente_por_insumo
                or created_at > mas_reciente_por_insumo[insumo_id]
            ):
                mas_reciente_por_insumo[insumo_id] = created_at
                costo_unitario_por_insumo[insumo_id] = float(dl['costo_unitario'])

    # 4) Porcentaje de merma por insumo (CU22, FICHA_TECNICA)
    merma_por_insumo = {}
    if insumo_ids:
        fichas = supabase.table('ficha_tecnica') \
            .select('insumo_id, porcentaje_merma') \
            .in_('insumo_id', insumo_ids) \
            .execute().data or []
        for f in fichas:
            if f.get('porcentaje_merma') is not None:
                merma_por_insumo[f['insumo_id']] = float(f['porcentaje_merma'])

    # 5) Nombres y precio de venta de los platos
    platos_response = supabase.table('plato') \
        .select('id, nombre, costo') \
        .in_('id', plato_ids) \
        .execute().data or []
    platos_info = {p['id']: p for p in platos_response}

    # 6) Agrupar detalles por plato y calcular costo teórico / real
    detalles_por_plato = {}
    for d in detalles:
        plato_id_d = receta_a_plato.get(d['receta_id'])
        detalles_por_plato.setdefault(plato_id_d, []).append(d)

    reporte = []
    for p_id, sus_detalles in detalles_por_plato.items():
        costo_teorico = 0.0
        costo_real = 0.0
        costos_incompletos = False

        for d in sus_detalles:
            insumo_id = d['insumo_id']
            cantidad = float(d['cantidad'])
            costo_unitario = costo_unitario_por_insumo.get(insumo_id)

            if costo_unitario is None:
                costos_incompletos = True
                costo_unitario = 0.0

            merma = merma_por_insumo.get(insumo_id, 0.0)

            costo_teorico += cantidad * costo_unitario
            costo_real += cantidad * costo_unitario * (1 + merma / 100)

        plato_info = platos_info.get(p_id, {})
        precio_venta = plato_info.get('costo')
        precio_venta = float(precio_venta) if precio_venta is not None else None

        margen = None
        if precio_venta:
            margen = round(((precio_venta - costo_real) / precio_venta) * 100, 2)

        reporte.append({
            'plato_id': p_id,
            'plato_nombre': plato_info.get('nombre', 'Desconocido'),
            'costo_teorico': round(costo_teorico, 2),
            'costo_real': round(costo_real, 2),
            'precio_venta': precio_venta,
            'margen': margen,
            'costos_incompletos': costos_incompletos,
        })

    reporte.sort(key=lambda x: x['plato_nombre'])
    return reporte


class ReporteCostosView(APIView):
    """
    Endpoint para obtener el reporte de costos en formato JSON.

    Método: GET
    URL: /api/reportes/costos-plato/
    Query params:
        - plato_id (opcional): si se indica, filtra el reporte a un solo plato

    Respuesta exitosa (200):
    {
        "reporte": [ {...}, ... ],
        "plato_mas_rentable": {...} | null,
        "plato_menos_rentable": {...} | null
    }
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        plato_id = request.query_params.get('plato_id')
        plato_id = int(plato_id) if plato_id else None

        try:
            supabase = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
            reporte = _calcular_reporte_costos(supabase, plato_id=plato_id)

            con_margen = [r for r in reporte if r['margen'] is not None]
            plato_mas_rentable = max(con_margen, key=lambda r: r['margen']) if con_margen else None
            plato_menos_rentable = min(con_margen, key=lambda r: r['margen']) if con_margen else None

            return Response({
                'reporte': reporte,
                'plato_mas_rentable': plato_mas_rentable,
                'plato_menos_rentable': plato_menos_rentable,
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error generando reporte de costos: {str(e)}")
            return Response(
                {'error': f'Error al generar el reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ReporteCostosPDFView(APIView):
    """
    Endpoint para descargar el reporte de costos en PDF.

    Método: GET
    URL: /api/reportes/costos-plato/pdf/
    Query params:
        - plato_id (opcional)

    Respuesta: archivo PDF (application/pdf) para descarga directa.
    Registra GENERAR_REPORTE_COSTOS en bitácora con formato='pdf'.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        plato_id = request.query_params.get('plato_id')
        plato_id = int(plato_id) if plato_id else None

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
            from reportlab.lib.styles import getSampleStyleSheet

            supabase = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
            reporte = _calcular_reporte_costos(supabase, plato_id=plato_id)

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer, pagesize=letter,
                topMargin=2 * cm, bottomMargin=2 * cm
            )
            styles = getSampleStyleSheet()
            elementos = []

            elementos.append(Paragraph(
                "Reporte de Costos por Plato", styles['Title']
            ))
            elementos.append(Paragraph(
                "Sistema de Gestión de Almacenes Gastronómicos · ODAA Simplificado",
                styles['Normal']
            ))
            elementos.append(Spacer(1, 0.5 * cm))

            data = [['Plato', 'Costo Teórico (Bs.)', 'Costo Real (Bs.)', 'Precio Venta (Bs.)', 'Margen (%)']]
            for r in reporte:
                data.append([
                    r['plato_nombre'],
                    f"{r['costo_teorico']:.2f}",
                    f"{r['costo_real']:.2f}",
                    f"{r['precio_venta']:.2f}" if r['precio_venta'] is not None else "N/D",
                    f"{r['margen']:.2f}%" if r['margen'] is not None else "N/D",
                ])

            if len(data) == 1:
                data.append(['Sin datos disponibles', '-', '-', '-', '-'])

            tabla = Table(data, colWidths=[5 * cm, 3.2 * cm, 3.2 * cm, 3.2 * cm, 2.8 * cm])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f97316')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elementos.append(tabla)

            doc.build(elementos)
            buffer.seek(0)

            ip_cliente = obtener_ip_cliente(request)
            registrar_accion(
                usuario_id=str(request.user.id),
                usuario_email=request.user.email,
                accion="GENERAR_REPORTE_COSTOS",
                detalles={"ip": ip_cliente, "formato": "pdf", "plato_id": plato_id}
            )

            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="reporte_costos_plato.pdf"'
            return response

        except Exception as e:
            logger.error(f"Error generando PDF de reporte de costos: {str(e)}")
            return Response(
                {'error': f'Error al generar el PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ReporteCostosExcelView(APIView):
    """
    Endpoint para descargar el reporte de costos en Excel.

    Método: GET
    URL: /api/reportes/costos-plato/excel/
    Query params:
        - plato_id (opcional)

    Respuesta: archivo .xlsx para descarga directa.
    Registra GENERAR_REPORTE_COSTOS en bitácora con formato='excel'.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        plato_id = request.query_params.get('plato_id')
        plato_id = int(plato_id) if plato_id else None

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            supabase = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
            reporte = _calcular_reporte_costos(supabase, plato_id=plato_id)

            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte de Costos"

            headers = ['Plato', 'Costo Teórico (Bs.)', 'Costo Real (Bs.)', 'Precio Venta (Bs.)', 'Margen (%)']
            ws.append(headers)

            header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col_num, _ in enumerate(headers, start=1):
                celda = ws.cell(row=1, column=col_num)
                celda.fill = header_fill
                celda.font = header_font
                celda.alignment = Alignment(horizontal="center")

            for r in reporte:
                ws.append([
                    r['plato_nombre'],
                    r['costo_teorico'],
                    r['costo_real'],
                    r['precio_venta'] if r['precio_venta'] is not None else "N/D",
                    r['margen'] if r['margen'] is not None else "N/D",
                ])

            for col_num, header in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(col_num)].width = max(18, len(header) + 4)

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            ip_cliente = obtener_ip_cliente(request)
            registrar_accion(
                usuario_id=str(request.user.id),
                usuario_email=request.user.email,
                accion="GENERAR_REPORTE_COSTOS",
                detalles={"ip": ip_cliente, "formato": "excel", "plato_id": plato_id}
            )

            response = HttpResponse(
                buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="reporte_costos_plato.xlsx"'
            return response

        except Exception as e:
            logger.error(f"Error generando Excel de reporte de costos: {str(e)}")
            return Response(
                {'error': f'Error al generar el Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )