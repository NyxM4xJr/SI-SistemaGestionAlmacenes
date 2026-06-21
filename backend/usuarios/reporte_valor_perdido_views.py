"""
============================================================
ARCHIVO: backend/usuarios/reporte_valor_perdido_views.py
CASO DE USO: CU25 - Generar Reporte de Valor Perdido
CICLO: 4
AUTOR: Mateo Hurtado
FECHA: 21/06/26
============================================================

DESCRIPCIÓN:
Calcula el reporte de pérdidas de inventario a partir de los
movimientos tipo 'merma' ya registrados en MOVIMIENTO_INVENTARIO
(CU14). Agrupa el valor perdido por causa, por período (día/semana/
mes) y por insumo, dentro de un rango de fechas opcional.

No se crea ninguna tabla nueva. Se generan archivos PDF (reportlab)
y Excel (openpyxl) reales para descarga. Solo la exportación deja
constancia en la bitácora (GENERAR_REPORTE_VALOR_PERDIDO) — la
consulta en JSON es de solo lectura y no se audita, mismo criterio
aplicado en ReporteCostosView (CU27).

Tablas consultadas (todas ya existentes):
- MOVIMIENTO_INVENTARIO  (tipo='merma', causa, valor_perdido, fecha_mov, insumo_id FK -> INSUMO)
- INSUMO                 (para mostrar el nombre del insumo)

Correspondencia con el diagrama de secuencia (CICLO4_DIAGRAMS_SPEC_MATEO.md):
- F1 alt    [rango de fechas válido / inválido]   -> validación antes de consultar
- F2 loop   [por cada movimiento de merma]         -> agrupación en memoria
- F3 alt    [formato pdf / excel]                  -> 2 vistas separadas (PDF / Excel)
- F4 critical                                      -> registrar_accion() atómica, solo en exportación
"""

import io
import logging

from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from supabase import create_client
from django.conf import settings
from django.http import HttpResponse

from bitacora.utils import registrar_accion, obtener_ip_cliente

logger = logging.getLogger(__name__)


# Categorías fijas para agrupar el campo `causa` (texto libre en la BD).
# Cualquier valor que no calce exactamente con las claves de abajo cae
# en "Otro" (no se descarta ninguna fila, no se bloquea el reporte por
# causa desconocida). Mapeo decidido en sesión de diseño (Paso 1).
_CATEGORIAS_CAUSA = {
    'Vencimiento': 'Vencimiento',
    'Error de manipulación': 'Mala manipulación',
    'Contaminación': 'Mala manipulación',
    'Daño físico': 'Producto dañado',
    'Deterioro por temperatura': 'Producto dañado',
}
_CAUSA_OTRO = 'Otro'


def _categoria_causa(causa: str) -> str:
    """Resuelve la categoría visual de una causa de merma (texto libre)."""
    if not causa:
        return _CAUSA_OTRO
    return _CATEGORIAS_CAUSA.get(causa.strip(), _CAUSA_OTRO)


def _clave_periodo(fecha_mov: str, agrupar_por: str) -> str:
    """
    Construye la clave de agrupación por período a partir del string
    ISO de fecha_mov, SIN pasar por datetime/new Date(), para evitar
    el bug de desfase UTC ya documentado en CU14/CU15.

    fecha_mov llega como 'YYYY-MM-DD...' (string ISO de Supabase).
    """
    fecha = fecha_mov[:10]  # 'YYYY-MM-DD'
    if agrupar_por == 'dia':
        return fecha
    if agrupar_por == 'semana':
        # Semana ISO calculada a partir del string, sin pasar por datetime.date
        # (se permite el único uso puntual de datetime aquí porque ISO week
        # no tiene una forma simple de calcularse por slicing de string).
        from datetime import date
        anio, mes, dia = (int(x) for x in fecha.split('-'))
        semana_iso = date(anio, mes, dia).isocalendar()[1]
        return f"{anio}-W{semana_iso:02d}"
    # 'mes' por defecto
    return fecha[:7]  # 'YYYY-MM'


def _calcular_reporte_valor_perdido(
    supabase,
    fecha_desde=None,
    fecha_hasta=None,
    insumo_id=None,
    agrupar_por='mes',
):
    """
    Función compartida por los 3 endpoints (JSON, PDF, Excel).

    Trae los movimientos tipo 'merma' dentro del rango/filtros
    indicados, resuelve el nombre del insumo (FK formal a INSUMO)
    y agrupa el valor perdido por causa, por período y por insumo.

    Retorna un dict:
    {
        "movimientos": [ {...}, ... ],       # filas crudas, con nombre de insumo
        "por_causa": [ {"causa": str, "valor_perdido": float, "cantidad_eventos": int}, ... ],
        "por_periodo": [ {"periodo": str, "valor_perdido": float}, ... ],   # ordenado asc
        "por_insumo": [ {"insumo_id": int, "insumo_nombre": str, "valor_perdido": float}, ... ],  # desc, top 5
        "valor_perdido_total": float,
        "total_eventos": int,
    }
    """
    # 1) Movimientos tipo merma, con filtros opcionales
    query = supabase.table('movimiento_inventario') \
        .select('id, insumo_id, fecha_mov, causa, valor_perdido') \
        .eq('tipo', 'merma')

    if fecha_desde:
        query = query.gte('fecha_mov', fecha_desde)
    if fecha_hasta:
        query = query.lte('fecha_mov', fecha_hasta)
    if insumo_id is not None:
        query = query.eq('insumo_id', insumo_id)

    movimientos = query.execute().data or []

    if not movimientos:
        return {
            'movimientos': [],
            'por_causa': [],
            'por_periodo': [],
            'por_insumo': [],
            'valor_perdido_total': 0.0,
            'total_eventos': 0,
        }

    # 2) Nombres de insumo (JOIN formal, FK declarada en movimiento_inventario.insumo_id)
    insumo_ids = list({m['insumo_id'] for m in movimientos if m.get('insumo_id') is not None})
    insumos_info = {}
    if insumo_ids:
        insumos_response = supabase.table('insumo') \
            .select('id, nombre') \
            .in_('id', insumo_ids) \
            .execute().data or []
        insumos_info = {i['id']: i['nombre'] for i in insumos_response}

    # 3) Agrupación — F2: loop por cada movimiento de merma
    acumulado_causa = {}       # categoria -> {valor_perdido, cantidad_eventos}
    acumulado_periodo = {}     # periodo -> valor_perdido
    acumulado_insumo = {}      # insumo_id -> {insumo_nombre, valor_perdido}
    valor_perdido_total = 0.0
    movimientos_detalle = []

    for m in movimientos:
        valor = float(m.get('valor_perdido') or 0)
        causa_original = m.get('causa')
        categoria = _categoria_causa(causa_original)
        periodo = _clave_periodo(m['fecha_mov'], agrupar_por)
        insumo_id_m = m.get('insumo_id')
        insumo_nombre = insumos_info.get(insumo_id_m, 'Desconocido')

        valor_perdido_total += valor

        # Por causa
        bucket_causa = acumulado_causa.setdefault(
            categoria, {'valor_perdido': 0.0, 'cantidad_eventos': 0}
        )
        bucket_causa['valor_perdido'] += valor
        bucket_causa['cantidad_eventos'] += 1

        # Por período
        acumulado_periodo[periodo] = acumulado_periodo.get(periodo, 0.0) + valor

        # Por insumo
        bucket_insumo = acumulado_insumo.setdefault(
            insumo_id_m, {'insumo_nombre': insumo_nombre, 'valor_perdido': 0.0}
        )
        bucket_insumo['valor_perdido'] += valor

        movimientos_detalle.append({
            'id': m['id'],
            'insumo_id': insumo_id_m,
            'insumo_nombre': insumo_nombre,
            'fecha_mov': m['fecha_mov'],
            'causa': causa_original,
            'categoria_causa': categoria,
            'valor_perdido': round(valor, 2),
        })

    por_causa = [
        {
            'causa': causa,
            'valor_perdido': round(datos['valor_perdido'], 2),
            'cantidad_eventos': datos['cantidad_eventos'],
        }
        for causa, datos in acumulado_causa.items()
    ]
    por_causa.sort(key=lambda x: x['valor_perdido'], reverse=True)

    por_periodo = [
        {'periodo': periodo, 'valor_perdido': round(valor, 2)}
        for periodo, valor in acumulado_periodo.items()
    ]
    por_periodo.sort(key=lambda x: x['periodo'])  # asc, para el gráfico de tendencia

    por_insumo = [
        {
            'insumo_id': insumo_id_i,
            'insumo_nombre': datos['insumo_nombre'],
            'valor_perdido': round(datos['valor_perdido'], 2),
        }
        for insumo_id_i, datos in acumulado_insumo.items()
    ]
    por_insumo.sort(key=lambda x: x['valor_perdido'], reverse=True)
    top_5_insumos = por_insumo[:5]

    movimientos_detalle.sort(key=lambda x: x['fecha_mov'], reverse=True)

    return {
        'movimientos': movimientos_detalle,
        'por_causa': por_causa,
        'por_periodo': por_periodo,
        'por_insumo': top_5_insumos,
        'valor_perdido_total': round(valor_perdido_total, 2),
        'total_eventos': len(movimientos),
    }


def _validar_y_extraer_filtros(request):
    """
    F1 — alt [rango de fechas válido / inválido].
    Extrae y valida los query params comunes a los 3 endpoints.
    Lanza ValueError con un mensaje de usuario si el rango es inválido.
    """
    fecha_desde = request.query_params.get('fecha_desde')
    fecha_hasta = request.query_params.get('fecha_hasta')
    insumo_id = request.query_params.get('insumo_id')
    agrupar_por = request.query_params.get('agrupar_por', 'mes')

    if agrupar_por not in ('dia', 'semana', 'mes'):
        raise ValueError("agrupar_por debe ser 'dia', 'semana' o 'mes'")

    if fecha_desde and fecha_hasta and fecha_desde > fecha_hasta:
        raise ValueError("fecha_desde no puede ser posterior a fecha_hasta")

    insumo_id = int(insumo_id) if insumo_id else None

    return fecha_desde, fecha_hasta, insumo_id, agrupar_por


class ReporteValorPerdidoView(APIView):
    """
    Endpoint para obtener el reporte de valor perdido en formato JSON.

    Método: GET
    URL: /api/reportes/valor-perdido/
    Query params (todos opcionales):
        - fecha_desde (YYYY-MM-DD)
        - fecha_hasta (YYYY-MM-DD)
        - insumo_id
        - agrupar_por: 'dia' | 'semana' | 'mes' (default 'mes')

    Respuesta exitosa (200):
    {
        "movimientos": [...],
        "por_causa": [...],
        "por_periodo": [...],
        "por_insumo": [...],
        "valor_perdido_total": float,
        "total_eventos": int
    }

    No registra bitácora (solo lectura) — mismo criterio que
    ReporteCostosView (CU27).
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            fecha_desde, fecha_hasta, insumo_id, agrupar_por = _validar_y_extraer_filtros(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            supabase = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
            reporte = _calcular_reporte_valor_perdido(
                supabase,
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                insumo_id=insumo_id,
                agrupar_por=agrupar_por,
            )
            return Response(reporte, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Error generando reporte de valor perdido: {str(e)}")
            return Response(
                {'error': f'Error al generar el reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ReporteValorPerdidoPDFView(APIView):
    """
    Endpoint para descargar el reporte de valor perdido en PDF.

    Método: GET
    URL: /api/reportes/valor-perdido/pdf/
    Query params: iguales a ReporteValorPerdidoView

    Respuesta: archivo PDF (application/pdf) para descarga directa.
    Registra GENERAR_REPORTE_VALOR_PERDIDO en bitácora con formato='pdf'.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            fecha_desde, fecha_hasta, insumo_id, agrupar_por = _validar_y_extraer_filtros(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
            from reportlab.lib.styles import getSampleStyleSheet

            supabase = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
            reporte = _calcular_reporte_valor_perdido(
                supabase,
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                insumo_id=insumo_id,
                agrupar_por=agrupar_por,
            )

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer, pagesize=letter,
                topMargin=2 * cm, bottomMargin=2 * cm
            )
            styles = getSampleStyleSheet()
            elementos = []

            elementos.append(Paragraph(
                "Reporte de Valor Perdido", styles['Title']
            ))
            elementos.append(Paragraph(
                "Sistema de Gestión de Almacenes Gastronómicos · ODAA Simplificado",
                styles['Normal']
            ))
            elementos.append(Spacer(1, 0.3 * cm))
            elementos.append(Paragraph(
                f"Valor perdido total: Bs. {reporte['valor_perdido_total']:.2f}"
                f"  ·  Eventos de merma: {reporte['total_eventos']}",
                styles['Normal']
            ))
            elementos.append(Spacer(1, 0.5 * cm))

            # F2 (loop) — Tabla "Por causa"
            elementos.append(Paragraph("Pérdidas por Causa", styles['Heading2']))
            data_causa = [['Causa', 'Valor Perdido (Bs.)', 'N° Eventos']]
            for r in reporte['por_causa']:
                data_causa.append([
                    r['causa'],
                    f"{r['valor_perdido']:.2f}",
                    str(r['cantidad_eventos']),
                ])
            if len(data_causa) == 1:
                data_causa.append(['Sin datos disponibles', '-', '-'])

            tabla_causa = Table(data_causa, colWidths=[7 * cm, 4 * cm, 3 * cm])
            tabla_causa.setStyle(_estilo_tabla_reporte())
            elementos.append(tabla_causa)
            elementos.append(Spacer(1, 0.5 * cm))

            # Tabla "Top 5 insumos"
            elementos.append(Paragraph("Top 5 Insumos con Mayor Pérdida", styles['Heading2']))
            data_insumo = [['Insumo', 'Valor Perdido (Bs.)']]
            for r in reporte['por_insumo']:
                data_insumo.append([
                    r['insumo_nombre'],
                    f"{r['valor_perdido']:.2f}",
                ])
            if len(data_insumo) == 1:
                data_insumo.append(['Sin datos disponibles', '-'])

            tabla_insumo = Table(data_insumo, colWidths=[9 * cm, 5 * cm])
            tabla_insumo.setStyle(_estilo_tabla_reporte())
            elementos.append(tabla_insumo)

            doc.build(elementos)
            buffer.seek(0)

            # F4 — critical: registrar bitácora (única llamada atómica, solo en exportación)
            ip_cliente = obtener_ip_cliente(request)
            registrar_accion(
                usuario_id=str(request.user.id),
                usuario_email=request.user.email,
                accion="GENERAR_REPORTE_VALOR_PERDIDO",
                detalles={
                    "ip": ip_cliente,
                    "formato": "pdf",
                    "fecha_desde": fecha_desde,
                    "fecha_hasta": fecha_hasta,
                    "insumo_id": insumo_id,
                }
            )

            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="reporte_valor_perdido.pdf"'
            return response

        except Exception as e:
            logger.error(f"Error generando PDF de reporte de valor perdido: {str(e)}")
            return Response(
                {'error': f'Error al generar el PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ReporteValorPerdidoExcelView(APIView):
    """
    Endpoint para descargar el reporte de valor perdido en Excel.

    Método: GET
    URL: /api/reportes/valor-perdido/excel/
    Query params: iguales a ReporteValorPerdidoView

    Respuesta: archivo .xlsx para descarga directa.
    Registra GENERAR_REPORTE_VALOR_PERDIDO en bitácora con formato='excel'.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            fecha_desde, fecha_hasta, insumo_id, agrupar_por = _validar_y_extraer_filtros(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            supabase = create_client(settings.SUPABASE_URL, settings.SUPABASE_KEY)
            reporte = _calcular_reporte_valor_perdido(
                supabase,
                fecha_desde=fecha_desde,
                fecha_hasta=fecha_hasta,
                insumo_id=insumo_id,
                agrupar_por=agrupar_por,
            )

            wb = Workbook()

            # Hoja 1: Por causa
            ws_causa = wb.active
            ws_causa.title = "Por Causa"
            _escribir_hoja(
                ws_causa,
                headers=['Causa', 'Valor Perdido (Bs.)', 'N° Eventos'],
                filas=[
                    [r['causa'], r['valor_perdido'], r['cantidad_eventos']]
                    for r in reporte['por_causa']
                ],
            )

            # Hoja 2: Por período
            ws_periodo = wb.create_sheet("Por Período")
            _escribir_hoja(
                ws_periodo,
                headers=['Período', 'Valor Perdido (Bs.)'],
                filas=[
                    [r['periodo'], r['valor_perdido']]
                    for r in reporte['por_periodo']
                ],
            )

            # Hoja 3: Top insumos
            ws_insumo = wb.create_sheet("Top Insumos")
            _escribir_hoja(
                ws_insumo,
                headers=['Insumo', 'Valor Perdido (Bs.)'],
                filas=[
                    [r['insumo_nombre'], r['valor_perdido']]
                    for r in reporte['por_insumo']
                ],
            )

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            # F4 — critical: registrar bitácora (única llamada atómica, solo en exportación)
            ip_cliente = obtener_ip_cliente(request)
            registrar_accion(
                usuario_id=str(request.user.id),
                usuario_email=request.user.email,
                accion="GENERAR_REPORTE_VALOR_PERDIDO",
                detalles={
                    "ip": ip_cliente,
                    "formato": "excel",
                    "fecha_desde": fecha_desde,
                    "fecha_hasta": fecha_hasta,
                    "insumo_id": insumo_id,
                }
            )

            response = HttpResponse(
                buffer.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="reporte_valor_perdido.xlsx"'
            return response

        except Exception as e:
            logger.error(f"Error generando Excel de reporte de valor perdido: {str(e)}")
            return Response(
                {'error': f'Error al generar el Excel: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


def _estilo_tabla_reporte():
    """
    Estilo reportlab compartido por las tablas del PDF.
    Mismo esquema visual que ReporteCostosPDFView (CU27): header
    naranja, filas alternadas, para mantener identidad visual entre
    los reportes del paquete Reportes y Análisis.
    """
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f97316')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ])


def _escribir_hoja(ws, headers, filas):
    """
    Helper para escribir una hoja de Excel con el mismo estilo de
    header naranja que ReporteCostosExcelView (CU27).
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws.append(headers)

    header_fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_num, _ in enumerate(headers, start=1):
        celda = ws.cell(row=1, column=col_num)
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = Alignment(horizontal="center")

    if not filas:
        ws.append(['Sin datos disponibles'] + ['-'] * (len(headers) - 1))
    else:
        for fila in filas:
            ws.append(fila)

    for col_num, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_num)].width = max(18, len(header) + 4)